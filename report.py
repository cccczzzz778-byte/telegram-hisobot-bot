from __future__ import annotations

from datetime import date, datetime, time
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from roster import Institution


HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
GREEN_FILL = PatternFill("solid", fgColor="00B050")
YELLOW_FILL = PatternFill("solid", fgColor="FFFF00")
RED_FILL = PatternFill("solid", fgColor="C00000")
WHITE_FONT = Font(color="FFFFFF", bold=True)
THIN = Side(style="thin", color="000000")


def create_report(roster: list[Institution], submissions: dict[int, dict], report_date: date,
                  deadline: time, title: str, output: Path) -> Path:
    wb = Workbook()
    ws = wb.active
    ws.title = report_date.strftime("%d.%m.%Y")
    headers = [
        "№", "Muassasa nomi", "Bosh ginekolog F.I.Sh.", "Telefon raqami",
        "Tug‘ruqxona boshlig‘i F.I.Sh.", "Telefon raqami",
        "Bosh doya F.I.Sh.", "Telefon raqami", title,
    ]
    ws.append(headers)
    for col, value in enumerate(headers, 1):
        cell = ws.cell(1, col, value)
        cell.fill, cell.font = HEADER_FILL, WHITE_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for row_no, item in enumerate(roster, 2):
        submission = submissions.get(item.number)
        if submission:
            raw_sent = submission["submitted_at"]
            sent = raw_sent if isinstance(raw_sent, datetime) else datetime.fromisoformat(raw_sent)
            late = sent.time().replace(tzinfo=None) > deadline
            status = f"{'Kechikdi' if late else 'Topshirdi'} — {sent:%H:%M}"
            fill = YELLOW_FILL if late else GREEN_FILL
        else:
            status, fill = "Topshirmadi", RED_FILL
        values = [
            item.number, item.name, item.chief_gynecologist, item.chief_gynecologist_phone,
            item.maternity_head, item.maternity_head_phone,
            item.chief_midwife, item.chief_midwife_phone, status,
        ]
        ws.append(values)
        ws.cell(row_no, 9).fill = fill
        ws.cell(row_no, 9).font = Font(color="FFFFFF" if fill != YELLOW_FILL else "000000", bold=True)

    for row in ws.iter_rows():
        for cell in row:
            cell.border = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    widths = [7, 38, 30, 18, 30, 18, 30, 18, 24]
    for index, width in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(index)].width = width
    ws.row_dimensions[1].height = 35
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    ws.sheet_view.showGridLines = False
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToWidth = 1
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    output.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output)
    return output
