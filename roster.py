from __future__ import annotations

import base64
import binascii
import os
import re
from dataclasses import dataclass
from difflib import SequenceMatcher
from pathlib import Path

from openpyxl import load_workbook


@dataclass(frozen=True)
class Institution:
    number: int
    name: str
    chief_gynecologist: str
    chief_gynecologist_phone: str
    maternity_head: str
    maternity_head_phone: str
    chief_midwife: str
    chief_midwife_phone: str
    aliases: tuple[str, ...]


def ensure_roster_file(path: Path) -> None:
    """Railway siridan ro‘yxat faylini tiklaydi; mavjud faylga tegmaydi."""
    if path.exists():
        return
    encoded = os.getenv("ROSTER_XLSX_BASE64", "").strip()
    if not encoded:
        raise FileNotFoundError(
            f"Muassasalar ro‘yxati topilmadi: {path}. "
            "ROSTER_XLSX_BASE64 muhit o‘zgaruvchisini kiriting."
        )
    try:
        content = base64.b64decode(encoded, validate=True)
    except (ValueError, binascii.Error) as exc:
        raise RuntimeError("ROSTER_XLSX_BASE64 noto‘g‘ri formatda") from exc
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_bytes(content)


def normalize(text: str) -> str:
    text = text.casefold().replace("’", "'").replace("`", "'")
    replacements = {
        "а": "a", "б": "b", "в": "v", "г": "g", "ғ": "g", "д": "d",
        "е": "e", "ё": "yo", "ж": "j", "з": "z", "и": "i", "й": "y",
        "к": "k", "қ": "k", "л": "l", "м": "m", "н": "n", "о": "o",
        "п": "p", "р": "r", "с": "s", "т": "t", "у": "u", "ў": "o",
        "ф": "f", "х": "x", "ҳ": "x", "ц": "s", "ч": "ch", "ш": "sh",
        "ъ": "'", "ь": "", "э": "e", "ю": "yu", "я": "ya",
    }
    text = "".join(replacements.get(char, char) for char in text)
    return " ".join(re.sub(r"[^0-9a-zа-я' ]+", " ", text).split())


def load_roster(path: Path) -> list[Institution]:
    if not path.exists():
        raise FileNotFoundError(f"Muassasalar ro‘yxati topilmadi: {path}")
    sheet = load_workbook(path, read_only=True, data_only=True).active
    result: list[Institution] = []
    for row in sheet.iter_rows(min_row=2, values_only=True):
        if not row or not row[1]:
            continue
        aliases = tuple(x.strip() for x in str(row[8] or "").split(";") if x.strip()) if len(row) > 8 else ()
        result.append(Institution(
            number=int(row[0] or len(result) + 1),
            name=str(row[1]).strip(),
            chief_gynecologist=str(row[2] or "").strip(),
            chief_gynecologist_phone=str(row[3] or "").strip(),
            maternity_head=str(row[4] or "").strip() if len(row) > 4 else "",
            maternity_head_phone=str(row[5] or "").strip() if len(row) > 5 else "",
            chief_midwife=str(row[6] or "").strip() if len(row) > 6 else "",
            chief_midwife_phone=str(row[7] or "").strip() if len(row) > 7 else "",
            aliases=aliases,
        ))
    return result


def _score(caption: str, candidate: str) -> float:
    caption_n, candidate_n = normalize(caption), normalize(candidate)
    if not candidate_n:
        return 0.0
    if candidate_n in caption_n:
        return 1.0
    wanted = set(candidate_n.split())
    actual = set(caption_n.split())
    token_score = len(wanted & actual) / max(1, len(wanted))
    sequence_score = SequenceMatcher(None, caption_n[: max(250, len(candidate_n) * 3)], candidate_n).ratio()
    return max(token_score, sequence_score)


def find_institution(caption: str, roster: list[Institution], threshold: float) -> tuple[Institution | None, float]:
    best: Institution | None = None
    best_score = 0.0
    for institution in roster:
        score = max(_score(caption, name) for name in (institution.name, *institution.aliases))
        if score > best_score:
            best, best_score = institution, score
    return (best, best_score) if best_score >= threshold else (None, best_score)
